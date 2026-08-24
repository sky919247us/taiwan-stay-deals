# -*- coding: utf-8 -*-
"""產出前端要吃的資料檔，以及供下載的 Excel / CSV。

輸出：
  public/data/stays.json   全量（前端一次載完，之後查詢都在瀏覽器端）
  public/data/meta.json    更新時間、統計、縣市鄉鎮清單
  public/data/downloads/*  xlsx / csv
  CHANGELOG-data.md        與上一版比對的增減紀錄
"""
import os, csv, json, gzip, math, datetime, collections
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from config import (ROOT, RAW, CACHE, DATA, DOWNLOADS, CATEGORIES, CITY_ORDER,
                    ISLAND_CITIES, jload, jdump)
from manual import load_overrides

STAYS = os.path.join(DATA, "stays.json")
META = os.path.join(DATA, "meta.json")

FIELDS = ["id", "name", "city", "town", "kind", "classes", "stars", "taiwan_host",
          "address", "phone_raw", "phones", "website", "booking_urls",
          "price_low", "price_high", "weekday_price", "lat", "lng", "geo_source",
          "categories", "plans", "discounts", "channels", "flags", "period", "license",
          "services", "parking_spaces", "accessible_rooms", "capacity",
          "price_final", "price_src", "price_note", "price_room", "price_url",
          "g_rating", "g_reviews", "g_uri"]

CSV_COLS = ["序號", "旅宿名稱", "縣市", "鄉鎮市區", "業別", "電話", "地址",
            "優惠類別", "平日雙人房價", "價格來源", "官網定價", "折抵金額",
            "優惠期限", "方案內容", "緯度", "經度", "座標來源",
            "Google評分", "Google評論數", "官網", "旅宿網詳情"]

CAT_NAME = {c["key"]: c["name"] for c in CATEGORIES}
DETAIL = "https://www.taiwanstay.net.tw/TSA/web_page/TSA020200.jsp?hohi_id="


def to_rows(stays):
    out = []
    for i, r in enumerate(sorted(stays, key=lambda x: (CITY_ORDER.index(x["city"])
                                                       if x["city"] in CITY_ORDER else 99,
                                                       x.get("town", ""), x["name"])), 1):
        out.append({
            "序號": i, "旅宿名稱": r["name"], "縣市": r["city"], "鄉鎮市區": r.get("town", ""),
            "業別": r["kind"], "電話": r.get("phone_raw", ""), "地址": r.get("address", ""),
            "優惠類別": "、".join(CAT_NAME[c] for c in r["categories"]),
            "平日雙人房價": r.get("price_final") or "",
            "價格來源": {"manual": "人工查核", "operator": "業者自報",
                     "plan": "方案說明", "website": "業者官網"}.get(r.get("price_src"), ""),
            "官網定價": r.get("price_low") or "",
            "折抵金額": "、".join(str(d) for d in r.get("discounts", [])),
            "優惠期限": r.get("period", ""),
            "方案內容": "\n\n".join("【%s】%s" % (CAT_NAME[p["category"]], p["text"])
                                for p in r["plans"]),
            "緯度": r.get("lat", ""), "經度": r.get("lng", ""),
            "座標來源": {"opendata": "開放資料", "geocode": "地理編碼",
                     "township": "鄉鎮約略"}.get(r.get("geo_source"), ""),
            "Google評分": r.get("g_rating") or "",
            "Google評論數": r.get("g_reviews") or "",
            "官網": r.get("website", ""), "旅宿網詳情": DETAIL + r["id"],
        })
    return out


def write_downloads(rows):
    with open(os.path.join(DOWNLOADS, "taiwan-stay-deals.csv"), "w",
              encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLS)
        w.writeheader()
        w.writerows(rows)

    wb = Workbook()
    fill = PatternFill("solid", fgColor="1F6F5C")
    font = Font(color="FFFFFF", bold=True)
    widths = [6, 26, 8, 10, 10, 18, 34, 20, 12, 10, 10, 12, 12, 60, 10, 10, 10, 10, 12, 30, 46]

    def sheet(ws, data):
        ws.append(CSV_COLS)
        for c in ws[1]:
            c.fill, c.font = fill, font
            c.alignment = Alignment(horizontal="center")
        for r in data:
            ws.append([r[k] for k in CSV_COLS])
        for i, wd in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = wd
        for row in ws.iter_rows(min_row=2):
            row[CSV_COLS.index("方案內容")].alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    sheet(wb.active, rows)
    wb.active.title = "全部旅宿"
    for city in CITY_ORDER:
        d = [r for r in rows if r["縣市"] == city]
        if d:
            sheet(wb.create_sheet(city), d)
    wb.save(os.path.join(DOWNLOADS, "taiwan-stay-deals.xlsx"))


def diff_previous(stays):
    prev = jload(STAYS)
    if not prev:
        return None
    old = {s["id"]: s["name"] for s in prev["stays"]}
    new = {s["id"]: s["name"] for s in stays}
    added = [new[i] for i in new if i not in old]
    removed = [old[i] for i in old if i not in new]
    return {"added": added, "removed": removed}


def write_changelog(change, total):
    if not change or (not change["added"] and not change["removed"]):
        return
    path = os.path.join(ROOT, "CHANGELOG-data.md")
    today = datetime.date.today().isoformat()
    head = "## %s（共 %d 家）\n\n" % (today, total)
    body = ""
    if change["added"]:
        body += "**新增 %d 家**：%s\n\n" % (len(change["added"]), "、".join(change["added"][:80]))
    if change["removed"]:
        body += "**下架 %d 家**：%s\n\n" % (len(change["removed"]), "、".join(change["removed"][:80]))
    old = ""
    if os.path.exists(path):
        old = open(path, encoding="utf-8").read()
        old = old.replace("# 資料更新紀錄\n\n", "")
    with open(path, "w", encoding="utf-8") as f:
        f.write("# 資料更新紀錄\n\n" + head + body + old)


def meters(a1, o1, a2, o2):
    r, p = 6371000.0, math.pi / 180
    x = (math.sin((a2 - a1) * p / 2) ** 2 +
         math.cos(a1 * p) * math.cos(a2 * p) * math.sin((o2 - o1) * p / 2) ** 2)
    return 2 * r * math.asin(math.sqrt(x))


def main():
    blob = jload(os.path.join(RAW, "merged.json"))
    if not blob:
        raise SystemExit("找不到 raw/merged.json，請先跑 enrich.py 與 extract.py")
    stays = blob["stays"]

    web = jload(os.path.join(CACHE, "webprice.json"), {}) or {}
    manual = load_overrides()
    places = jload(os.path.join(CACHE, "places.json"), {}) or {}
    plan = jload(os.path.join(CACHE, "planprice.json"), {}) or {}
    moved = []

    for r in stays:
        for k in ("town", "address", "website", "license"):
            r.setdefault(k, "")
        for k in ("stars", "price_low", "price_high", "weekday_price"):
            r.setdefault(k, 0)
        r.setdefault("classes", [])
        r.setdefault("booking_urls", [])
        r.setdefault("taiwan_host", False)
        r.setdefault("geo_source", "")

        # 三種可信價格的優先序（開放資料的「定價」是牌價，永遠不列入）：
        #   1 人工查核  2 業者為活動自報  3 從業者官網抽取
        m, w = manual.get(r["id"]), web.get(r["id"]) or {}
        pt = plan.get(r["id"]) or {}
        if m:
            r["price_final"], r["price_src"] = m["price"], "manual"
            r["price_note"] = "、".join(x for x in (m["source"], m["date"], m["note"]) if x)
        elif r["weekday_price"]:
            r["price_final"], r["price_src"] = r["weekday_price"], "operator"
            r["price_note"] = "業者為本活動自報的平日雙人房價"
        # 只收模型確認是「平日價」的。標 unknown 的那批實測是大飯店官網的
        # 牌價（中位數 6,650、最高 16,000），跟政府那個灌水定價同一種東西，
        # 拿來當平日房價會再騙人一次。
        elif w.get("price") and w.get("basis") == "weekday":
            r["price_final"], r["price_src"] = w["price"], "website"
            r["price_note"] = "取自業者官網 %s：%s" % (w.get("fetched_at", ""),
                                                 w.get("evidence", ""))
            r["price_room"] = w.get("room", "")
            r["price_url"] = w.get("url", "")
        # 和官網抽價同一條規則：只收確認為平日房價的。標 unknown 的多半是
        # 「折抵補助後每晚只要 X」那種淨價，跟房價不是同一個東西。
        elif pt.get("price") and pt.get("basis") == "weekday":
            # 業者自己寫在方案說明裡的房價，來源就是官方名單本身，可信度高
            r["price_final"], r["price_src"] = pt["price"], "plan"
            r["price_note"] = "業者寫在方案說明中：" + (pt.get("evidence") or "")
            r["price_room"] = pt.get("room", "")
        else:
            r["price_final"], r["price_src"], r["price_note"] = 0, "", ""

        # Google 評分（Places API）。旅宿沒有 priceLevel，所以這裡只拿得到評價，
        # 但評分與評論數本身就是很有用的排序依據。
        g = places.get(r["id"]) or {}
        if g.get("place_id"):
            r["g_rating"] = g.get("rating") or 0
            r["g_reviews"] = g.get("reviews") or 0
            r["g_uri"] = g.get("maps_uri") or ""

            # 開放資料的座標有些明顯是錯的（6 家馬祖民宿被標在 24.7783,120.9881，
            # 那是台灣本島地理中心的預設值）。比對已通過縣市檢查，差距超過
            # 500 公尺就採信 Google 的座標。
            if g.get("lat") and r.get("lat"):
                d = meters(r["lat"], r["lng"], g["lat"], g["lng"])
                if d > 500:
                    r["lat"], r["lng"] = round(g["lat"], 6), round(g["lng"], 6)
                    r["geo_source"] = "google"
                    moved.append((r["name"], int(d)))

    slim = []
    for r in stays:
        s = {k: r.get(k) for k in FIELDS}
        # 瘦身：flags 只留為真的鍵；plans 去掉可由 key 推得的類別名稱與空連結
        s["flags"] = sorted(k for k, v in (r.get("flags") or {}).items() if v)
        s["plans"] = [{k: v for k, v in p.items()
                       if k in ("category", "text", "period") or (k == "links" and v)}
                      for p in r["plans"]]
        s = {k: v for k, v in s.items() if v not in ("", 0, [], False, None)}
        slim.append(s)
    change = diff_previous(slim)

    towns = collections.defaultdict(set)
    for r in slim:
        if r.get("town"):
            towns[r["city"]].add(r["town"])

    counts = collections.Counter()
    for r in slim:
        for c in r["categories"]:
            counts[c] += 1

    meta = {
        "updated_at": datetime.datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": "交通部觀光署臺灣旅宿網 優宿專區",
        "source_url": "https://www.taiwanstay.net.tw/TSA/web_page/TSA060100.jsp",
        "opendata_update": blob.get("opendata_update"),
        "total": len(slim),
        "categories": [{"key": c["key"], "name": c["name"], "count": counts[c["key"]]}
                       for c in CATEGORIES],
        "cities": [{"name": c, "count": sum(1 for r in slim if r["city"] == c),
                    "island": c in ISLAND_CITIES,
                    "towns": sorted(towns.get(c, []))}
                   for c in CITY_ORDER if any(r["city"] == c for r in slim)],
        "kinds": dict(collections.Counter(r["kind"] for r in slim)),
        "services": [{"name": n, "count": c} for n, c in
                     collections.Counter(x for r in slim for x in r.get("services", [])).most_common()],
        # 從最終資料重算，不要沿用 enrich 階段的數字：build 這裡還會用
        # Places 比中的座標覆蓋掉開放資料明顯偏移的那些。
        "geo_stat": dict(collections.Counter(
            r.get("geo_source") or "none" for r in slim)),
        "price_stat": dict(collections.Counter(
            r.get("price_src") or "none" for r in slim)),
        "rating_count": sum(1 for r in slim if r.get("g_rating")),
        "change": change,
    }

    jdump({"meta": {k: meta[k] for k in ("updated_at", "total", "source", "source_url")},
           "stays": slim}, STAYS, indent=None)
    jdump(meta, META)

    rows = to_rows(slim)
    write_downloads(rows)
    write_changelog(change, len(slim))

    if moved:
        moved.sort(key=lambda x: -x[1])
        print("[座標] 改用 Google 座標 %d 家，位移最大的：%s"
              % (len(moved), "、".join("%s %dm" % m for m in moved[:3])))
    size = os.path.getsize(STAYS)
    gz = len(gzip.compress(open(STAYS, "rb").read()))
    print("[輸出] stays.json %.0f KB（gzip 後 %.0f KB）" % (size / 1024, gz / 1024))
    print("[輸出] %d 家、%d 縣市、%s" % (len(slim), len(meta["cities"]), dict(counts)))
    if change:
        print("[異動] 新增 %d、下架 %d" % (len(change["added"]), len(change["removed"])))


if __name__ == "__main__":
    main()
